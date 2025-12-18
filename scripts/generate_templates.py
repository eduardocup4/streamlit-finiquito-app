"""
Generate Sample Excel Templates for Finiquito Documents
Creates template XLSX files for each document type
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

# Create templates directory
TEMPLATE_DIR = Path("/home/claude/finiquito_app/storage/templates")
TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)

def create_header_style():
    """Create standard header style"""
    return {
        'font': Font(bold=True, size=12),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'fill': PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    }

def create_border():
    """Create standard border"""
    thin = Side(style='thin')
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def create_f_finiquito_template():
    """Create template for Finiquito document"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Finiquito"
    
    # Company header
    ws.merge_cells('A1:H1')
    ws['A1'] = "ALIANZA SEGUROS S.A."
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:H2')
    ws['A2'] = "LIQUIDACIÓN DE BENEFICIOS SOCIALES"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Employee data section
    ws['A4'] = "DATOS DEL EMPLEADO"
    ws['A4'].font = Font(bold=True)
    
    ws['A5'] = "Nombre Completo:"
    ws['C5'] = "{{nombre}}"
    ws['E5'] = "CI:"
    ws['G5'] = "{{ci}}"
    
    ws['A6'] = "Cargo:"
    ws['C6'] = "{{cargo}}"
    ws['E6'] = "Empresa:"
    ws['G6'] = "{{empresa}}"
    
    ws['A7'] = "Fecha Ingreso:"
    ws['C7'] = "{{fecha_ingreso}}"
    ws['E7'] = "Fecha Retiro:"
    ws['G7'] = "{{fecha_retiro}}"
    
    # Calculation section
    ws['A9'] = "CÁLCULO DE BENEFICIOS"
    ws['A9'].font = Font(bold=True)
    
    headers = ['Concepto', 'Base', 'Tiempo', 'Factor', 'Monto']
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=i)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D0D0D0', end_color='D0D0D0', fill_type='solid')
        cell.border = create_border()
    
    # Benefits rows (will be filled dynamically)
    benefits_start = 11
    ws[f'A{benefits_start}'] = "{{beneficios}}"
    
    # Totals section
    ws['A20'] = "RESUMEN"
    ws['A20'].font = Font(bold=True)
    
    ws['A21'] = "Total Beneficios:"
    ws['E21'] = "{{total_beneficios}}"
    
    ws['A22'] = "Total Deducciones:"
    ws['E22'] = "{{total_deducciones}}"
    
    ws['A23'] = "NETO A PAGAR:"
    ws['A23'].font = Font(bold=True)
    ws['E23'] = "{{neto_pagar}}"
    ws['E23'].font = Font(bold=True)
    
    # Signature section
    ws['A26'] = "_" * 30
    ws['E26'] = "_" * 30
    ws['A27'] = "Firma Empleado"
    ws['E27'] = "Firma Autorizada"
    
    # Adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Save template
    template_path = TEMPLATE_DIR / "f_finiquito_template.xlsx"
    wb.save(template_path)
    print(f"Created: {template_path}")
    return template_path

def create_memo_finalizacion_template():
    """Create template for Memorandum de Finalización"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Memo"
    
    # Header
    ws['A1'] = "ALIANZA SEGUROS S.A."
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A3'] = "MEMORANDUM DE FINALIZACIÓN DE CONTRATO"
    ws['A3'].font = Font(bold=True, size=12)
    
    ws['A5'] = "CITE:"
    ws['B5'] = "{{cite}}"
    
    ws['A6'] = "FECHA:"
    ws['B6'] = "{{fecha}}"
    
    ws['A7'] = "DE:"
    ws['B7'] = "{{de}}"
    
    ws['A8'] = "PARA:"
    ws['B8'] = "{{para}}"
    
    ws['A9'] = "REF:"
    ws['B9'] = "Finalización de Contrato Laboral"
    
    # Body
    ws['A11'] = "Mediante el presente memorandum, se comunica la finalización del contrato laboral de:"
    ws.merge_cells('A11:F11')
    
    ws['A13'] = "Nombre:"
    ws['B13'] = "{{nombre}}"
    ws['D13'] = "CI:"
    ws['E13'] = "{{ci}}"
    
    ws['A14'] = "Cargo:"
    ws['B14'] = "{{cargo}}"
    ws['D14'] = "Unidad:"
    ws['E14'] = "{{unidad}}"
    
    ws['A16'] = "Fecha de Ingreso:"
    ws['B16'] = "{{fecha_ingreso}}"
    ws['D16'] = "Fecha de Retiro:"
    ws['E16'] = "{{fecha_retiro}}"
    
    ws['A18'] = "Motivo de Retiro:"
    ws['B18'] = "{{motivo_retiro}}"
    ws.merge_cells('B18:E18')
    
    ws['A20'] = "Se instruye proceder con:"
    ws['A21'] = "• Liquidación de beneficios sociales"
    ws['A22'] = "• Entrega de certificado de trabajo"
    ws['A23'] = "• Desafiliación de sistemas"
    
    # Signature
    ws['A26'] = "_" * 30
    ws['A27'] = "Gerente de Recursos Humanos"
    
    # Save template
    template_path = TEMPLATE_DIR / "memo_finalizacion_template.xlsx"
    wb.save(template_path)
    print(f"Created: {template_path}")
    return template_path

def create_f_salida_template():
    """Create template for Formulario de Salida"""
    wb = Workbook()
    ws = wb.active
    ws.title = "FormSalida"
    
    ws['A1'] = "FORMULARIO DE SALIDA DE PERSONAL"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Employee data
    ws['A3'] = "DATOS DEL EMPLEADO"
    ws['A3'].font = Font(bold=True)
    
    ws['A4'] = "Nombre:"
    ws['B4'] = "{{nombre}}"
    ws['D4'] = "CI:"
    ws['E4'] = "{{ci}}"
    
    ws['A5'] = "Cargo:"
    ws['B5'] = "{{cargo}}"
    ws['D5'] = "Empresa:"
    ws['E5'] = "{{empresa}}"
    
    # Checklist
    ws['A7'] = "CHECKLIST DE SALIDA"
    ws['A7'].font = Font(bold=True)
    
    checklist_items = [
        "Entrega de credencial",
        "Entrega de llaves",
        "Entrega de equipos informáticos",
        "Entrega de uniformes",
        "Cierre de cuentas de correo",
        "Desactivación de accesos",
        "Entrega de documentación pendiente",
        "Firma de finiquito"
    ]
    
    ws['A8'] = "Item"
    ws['B8'] = "Completo"
    ws['C8'] = "Pendiente"
    ws['D8'] = "N/A"
    ws['E8'] = "Observaciones"
    
    for i, item in enumerate(checklist_items, 9):
        ws[f'A{i}'] = item
        ws[f'B{i}'] = "[ ]"
        ws[f'C{i}'] = "[ ]"
        ws[f'D{i}'] = "[ ]"
        ws[f'E{i}'] = ""
    
    # Signatures
    ws['A19'] = "_" * 25
    ws['C19'] = "_" * 25
    ws['E19'] = "_" * 25
    ws['A20'] = "Empleado"
    ws['C20'] = "RRHH"
    ws['E20'] = "Supervisor"
    
    # Save template
    template_path = TEMPLATE_DIR / "f_salida_template.xlsx"
    wb.save(template_path)
    print(f"Created: {template_path}")
    return template_path

def create_f_equipos_template():
    """Create template for Formulario de Equipos"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipos"
    
    ws['A1'] = "FORMULARIO DE ENTREGA DE EQUIPOS"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Employee data
    ws['A3'] = "Empleado:"
    ws['B3'] = "{{nombre}}"
    ws['D3'] = "CI:"
    ws['E3'] = "{{ci}}"
    
    ws['A4'] = "Área:"
    ws['B4'] = "{{area}}"
    ws['D4'] = "Fecha:"
    ws['E4'] = "{{fecha}}"
    
    # Equipment list
    ws['A6'] = "LISTADO DE EQUIPOS"
    ws['A6'].font = Font(bold=True)
    
    headers = ['Item', 'Descripción', 'Código/Serie', 'Estado', 'Entregado', 'Observaciones']
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=i)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D0D0D0', end_color='D0D0D0', fill_type='solid')
    
    equipment_items = [
        "Laptop",
        "Mouse",
        "Teclado", 
        "Monitor",
        "Celular",
        "Tarjetas de acceso",
        "Otros"
    ]
    
    for i, item in enumerate(equipment_items, 8):
        ws[f'A{i}'] = item
        ws[f'B{i}'] = ""
        ws[f'C{i}'] = ""
        ws[f'D{i}'] = "[ ] Bueno [ ] Regular [ ] Malo"
        ws[f'E{i}'] = "[ ] Sí [ ] No"
        ws[f'F{i}'] = ""
    
    # Confirmation
    ws['A17'] = "Confirmo que he entregado todos los equipos listados en las condiciones indicadas."
    ws.merge_cells('A17:F17')
    
    # Signatures
    ws['A19'] = "_" * 30
    ws['D19'] = "_" * 30
    ws['A20'] = "Firma Empleado"
    ws['D20'] = "Firma IT/Activos"
    
    # Save template
    template_path = TEMPLATE_DIR / "f_equipos_template.xlsx"
    wb.save(template_path)
    print(f"Created: {template_path}")
    return template_path

def create_contable_preview_template():
    """Create template for Contable Preview"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Contable"
    
    ws['A1'] = "VISTA CONTABLE DE LIQUIDACIÓN"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A3'] = f"Fecha: {{fecha}}"
    ws['A4'] = f"Empleado: {{nombre}}"
    ws['A5'] = f"CI: {{ci}}"
    
    # Accounting table
    ws['A7'] = "ASIENTOS CONTABLES"
    ws['A7'].font = Font(bold=True)
    
    headers = ['Concepto', 'Debe', 'Haber', 'Referencia']
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=i)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='D0D0D0', end_color='D0D0D0', fill_type='solid')
        cell.border = create_border()
    
    # Sample entries (will be filled dynamically)
    ws['A9'] = "{{conceptos_contables}}"
    
    # Totals
    ws['A20'] = "TOTALES"
    ws['A20'].font = Font(bold=True)
    ws['B20'] = "{{total_debe}}"
    ws['C20'] = "{{total_haber}}"
    
    # Notes
    ws['A22'] = "NOTA: No se incluyen códigos de cuenta. Usar nombres de concepto únicamente."
    ws['A22'].font = Font(italic=True, size=10)
    ws.merge_cells('A22:F22')
    
    # Save template
    template_path = TEMPLATE_DIR / "contable_preview_template.xlsx"
    wb.save(template_path)
    print(f"Created: {template_path}")
    return template_path

def create_rechazo_post_template():
    """Create template for Rechazo Post-Examen"""
    wb = Workbook()
    ws = wb.active
    ws.title = "RechazoPost"
    
    ws['A1'] = "NOTIFICACIÓN DE RECHAZO POST-EXAMEN MÉDICO"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A3'] = "Fecha:"
    ws['B3'] = "{{fecha}}"
    
    ws['A5'] = "DATOS DEL EMPLEADO"
    ws['A5'].font = Font(bold=True)
    
    ws['A6'] = "Nombre:"
    ws['B6'] = "{{nombre}}"
    ws['D6'] = "CI:"
    ws['E6'] = "{{ci}}"
    
    ws['A7'] = "Cargo:"
    ws['B7'] = "{{cargo}}"
    ws['D7'] = "Empresa:"
    ws['E7'] = "{{empresa}}"
    
    ws['A9'] = "RESULTADO DEL EXAMEN"
    ws['A9'].font = Font(bold=True)
    
    ws['A10'] = "Fecha de Examen:"
    ws['B10'] = "{{fecha_examen}}"
    
    ws['A11'] = "Resultado:"
    ws['B11'] = "NO APTO"
    ws['B11'].font = Font(bold=True, color='FF0000')
    
    ws['A13'] = "En consecuencia, se procede con la finalización del proceso de contratación."
    ws.merge_cells('A13:E13')
    
    ws['A15'] = "LIQUIDACIÓN CORRESPONDIENTE"
    ws['A15'].font = Font(bold=True)
    
    ws['A16'] = "Días trabajados:"
    ws['B16'] = "{{dias_trabajados}}"
    
    ws['A17'] = "Monto a liquidar:"
    ws['B17'] = "{{monto_liquidar}}"
    
    # Signatures
    ws['A20'] = "_" * 30
    ws['D20'] = "_" * 30
    ws['A21'] = "RRHH"
    ws['D21'] = "Recibí Conforme"
    
    # Save template
    template_path = TEMPLATE_DIR / "rechazo_post_template.xlsx"
    wb.save(template_path)
    print(f"Created: {template_path}")
    return template_path

def main():
    """Generate all templates"""
    print("Generating Excel Templates...")
    print("=" * 50)
    
    templates = [
        create_f_finiquito_template(),
        create_memo_finalizacion_template(),
        create_f_salida_template(),
        create_f_equipos_template(),
        create_contable_preview_template(),
        create_rechazo_post_template()
    ]
    
    print("=" * 50)
    print(f"✅ Created {len(templates)} templates in {TEMPLATE_DIR}")
    
    # Create a summary file
    summary_path = TEMPLATE_DIR / "templates_info.txt"
    with open(summary_path, 'w') as f:
        f.write("Excel Templates for Finiquito System\n")
        f.write(f"Generated: {datetime.now()}\n")
        f.write("=" * 50 + "\n\n")
        
        for template in templates:
            f.write(f"- {template.name}\n")
        
        f.write("\n" + "=" * 50 + "\n")
        f.write("These templates use {{placeholders}} that will be replaced by the application.\n")
        f.write("Templates can be customized and uploaded through the Admin panel.\n")
    
    print(f"Summary saved to: {summary_path}")
    
    return templates

if __name__ == "__main__":
    main()
