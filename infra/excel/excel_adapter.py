"""
Excel adapter for reading payroll data and writing output documents
"""
import pandas as pd
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime, date
from decimal import Decimal
import os
import sys

sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

from domain.entities import Employee, PayrollMonth, FiniquitoCalculationResult
from config import field_config

class ExcelReader:
    """Reader for Excel payroll and RDP files"""
    
    def __init__(self):
        self.field_config = field_config
    
    def read_excel_file(
        self, 
        file_path: str, 
        sheet_name: Optional[str] = None
    ) -> pd.DataFrame:
        """
        Read an Excel file and return DataFrame
        """
        try:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            
            # Clean column names
            df.columns = [str(col).strip() for col in df.columns]
            
            return df
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        """
        Get all sheet names from an Excel file
        """
        try:
            xl_file = pd.ExcelFile(file_path)
            return xl_file.sheet_names
        except Exception as e:
            raise Exception(f"Error getting sheet names: {str(e)}")
    
    def normalize_column_name(self, column: str) -> str:
        """
        Normalize column name for matching
        """
        return column.lower().strip().replace(' ', '_').replace('.', '')
    
    def find_column_mapping(
        self, 
        df_columns: List[str], 
        required_fields: List[str]
    ) -> Dict[str, str]:
        """
        Auto-detect column mappings based on aliases
        """
        mapping = {}
        normalized_columns = {self.normalize_column_name(col): col for col in df_columns}
        
        for field in required_fields:
            # Check aliases
            if field in self.field_config.FIELD_ALIASES:
                aliases = self.field_config.FIELD_ALIASES[field]
                for alias in aliases:
                    normalized_alias = self.normalize_column_name(alias)
                    if normalized_alias in normalized_columns:
                        mapping[field] = normalized_columns[normalized_alias]
                        break
            
            # If not found, try exact match
            if field not in mapping:
                if field in df_columns:
                    mapping[field] = field
        
        return mapping
    
    def extract_employee_data(
        self,
        payroll_df: pd.DataFrame,
        rdp_df: pd.DataFrame,
        employee_ci: str,
        employee_empresa: str,
        payroll_mapping: Dict[str, str],
        rdp_mapping: Dict[str, str]
    ) -> Employee:
        """
        Extract employee data from payroll and RDP DataFrames
        """
        # Find employee in payroll (using most recent month)
        payroll_row = None
        for _, row in payroll_df.iterrows():
            if (str(row.get(payroll_mapping.get('ci', 'ci'), '')).strip() == employee_ci and 
                str(row.get(payroll_mapping.get('empresa', 'empresa'), '')).strip() == employee_empresa):
                payroll_row = row
                break
        
        if payroll_row is None:
            raise ValueError(f"Employee not found in payroll: {employee_ci} - {employee_empresa}")
        
        # Find employee in RDP
        rdp_row = None
        for _, row in rdp_df.iterrows():
            if (str(row.get(rdp_mapping.get('ci', 'ci'), '')).strip() == employee_ci and 
                str(row.get(rdp_mapping.get('empresa', 'empresa'), '')).strip() == employee_empresa):
                rdp_row = row
                break
        
        # Parse dates
        fecha_ingreso = pd.to_datetime(payroll_row[payroll_mapping['fecha_ingreso']]).date()
        fecha_nacimiento = pd.to_datetime(payroll_row[payroll_mapping['fecha_nacimiento']]).date()
        
        # Create Employee object
        return Employee(
            ci=employee_ci,
            name=str(payroll_row[payroll_mapping['nombre']]),
            empresa=employee_empresa,
            unidad=str(payroll_row[payroll_mapping['unidad']]),
            ocupacion=str(payroll_row[payroll_mapping['ocupacion']]),
            fecha_ingreso=fecha_ingreso,
            fecha_nacimiento=fecha_nacimiento,
            extension=str(rdp_row[rdp_mapping['extension']]) if rdp_row is not None else None,
            estado_civil=str(rdp_row[rdp_mapping['estado_civil']]) if rdp_row is not None else None,
            domicilio=str(rdp_row[rdp_mapping['domicilio']]) if rdp_row is not None else None
        )
    
    def extract_payroll_months(
        self,
        payroll_dfs: List[Tuple[pd.DataFrame, str]],  # [(df, month_name), ...]
        employee_ci: str,
        employee_empresa: str,
        mapping: Dict[str, str],
        include_otros_bonos: bool = False
    ) -> List[PayrollMonth]:
        """
        Extract payroll data for specific employee from 3 months
        """
        payroll_months = []
        
        for df, month_name in payroll_dfs:
            # Find employee row
            employee_row = None
            for _, row in df.iterrows():
                if (str(row.get(mapping.get('ci', 'ci'), '')).strip() == employee_ci and 
                    str(row.get(mapping.get('empresa', 'empresa'), '')).strip() == employee_empresa):
                    employee_row = row
                    break
            
            if employee_row is None:
                raise ValueError(f"Employee not found in {month_name}: {employee_ci} - {employee_empresa}")
            
            # Extract amounts
            haber_basico = Decimal(str(employee_row[mapping['haber_basico']]))
            bono_antiguedad = Decimal(str(employee_row[mapping['bono_antiguedad']]))
            total_ganado = Decimal(str(employee_row[mapping['total_ganado']]))
            
            otros_bonos = None
            if include_otros_bonos and 'otros_bonos' in mapping:
                otros_bonos = Decimal(str(employee_row.get(mapping['otros_bonos'], 0)))
            
            payroll_months.append(
                PayrollMonth(
                    month_name=month_name,
                    year_month=datetime.now().strftime("%Y-%m"),  # This should be extracted from data
                    haber_basico=haber_basico,
                    bono_antiguedad=bono_antiguedad,
                    otros_bonos=otros_bonos,
                    total_ganado=total_ganado
                )
            )
        
        return payroll_months

class ExcelWriter:
    """Writer for Excel output documents"""
    
    def create_finiquito_document(
        self,
        calculation_result: FiniquitoCalculationResult,
        template_path: Optional[str] = None,
        output_path: str = None
    ) -> str:
        """
        Create finiquito Excel document
        """
        if template_path and Path(template_path).exists():
            wb = load_workbook(template_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            self._create_finiquito_structure(ws)
        
        # Fill data
        self._fill_finiquito_data(ws, calculation_result)
        
        # Save
        if not output_path:
            output_path = f"finiquito_{calculation_result.employee.ci}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb.save(output_path)
        return output_path
    
    def _create_finiquito_structure(self, ws):
        """
        Create basic structure for finiquito document
        """
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = "LIQUIDACIÓN DE BENEFICIOS SOCIALES"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = [
            ('A3', 'DATOS DEL TRABAJADOR'),
            ('A10', 'CÁLCULO DE BENEFICIOS'),
            ('A25', 'RESUMEN')
        ]
        
        for cell, header in headers:
            ws[cell] = header
            ws[cell].font = Font(size=12, bold=True)
            ws[cell].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    def _fill_finiquito_data(self, ws, result: FiniquitoCalculationResult):
        """
        Fill finiquito document with calculation data
        """
        # Employee data
        employee_data = [
            ('A4', 'Nombre:', 'B4', result.employee.name),
            ('A5', 'CI:', 'B5', result.employee.ci),
            ('A6', 'Empresa:', 'B6', result.employee.empresa),
            ('A7', 'Fecha Ingreso:', 'B7', result.employee.fecha_ingreso.strftime('%d/%m/%Y')),
            ('A8', 'Fecha Salida:', 'B8', result.case_params.pay_until_date.strftime('%d/%m/%Y')),
            ('D4', 'Cargo:', 'E4', result.employee.ocupacion),
            ('D5', 'Unidad:', 'E5', result.employee.unidad),
            ('D6', 'Antigüedad:', 'E6', result.antiguedad.formatted),
            ('D7', 'Promedio Salarial:', 'E7', f"Bs. {result.salary_average:,.2f}"),
        ]
        
        for label_cell, label, value_cell, value in employee_data:
            ws[label_cell] = label
            ws[label_cell].font = Font(bold=True)
            ws[value_cell] = value
        
        # Benefits detail
        row = 11
        ws[f'A{row}'] = "Concepto"
        ws[f'D{row}'] = "Monto (Bs.)"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'D{row}'].font = Font(bold=True)
        
        row += 1
        for benefit in result.benefits:
            ws[f'A{row}'] = benefit.description
            ws[f'D{row}'] = f"{benefit.calculated_amount:,.2f}"
            row += 1
        
        # Deductions
        if result.deductions:
            row += 1
            ws[f'A{row}'] = "DEDUCCIONES"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for deduction in result.deductions:
                ws[f'A{row}'] = deduction.description
                ws[f'D{row}'] = f"{deduction.calculated_amount:,.2f}"
                row += 1
        
        # Summary
        row = 26
        summary_data = [
            ('Total Beneficios:', f"Bs. {result.total_benefits:,.2f}"),
            ('Total Deducciones:', f"Bs. {result.total_deductions:,.2f}"),
            ('NETO A PAGAR:', f"Bs. {result.net_payment:,.2f}"),
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'D{row}'] = value
            if 'NETO' in label:
                ws[f'A{row}'].font = Font(size=12, bold=True)
                ws[f'D{row}'].font = Font(size=12, bold=True)
            row += 1
        
        # Adjust column widths
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 15
    
    def create_memo_finalizacion(
        self,
        calculation_result: FiniquitoCalculationResult,
        template_path: Optional[str] = None,
        include_cite: bool = True,
        cite_number: Optional[str] = None,
        output_path: str = None
    ) -> str:
        """
        Create memo de finalización document
        """
        if template_path and Path(template_path).exists():
            wb = load_workbook(template_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            self._create_memo_structure(ws)
        
        # Fill data
        self._fill_memo_data(ws, calculation_result, include_cite, cite_number)
        
        # Save
        if not output_path:
            output_path = f"memo_{calculation_result.employee.ci}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb.save(output_path)
        return output_path
    
    def _create_memo_structure(self, ws):
        """
        Create structure for memo document
        """
        # Header
        ws.merge_cells('A1:F1')
        ws['A1'] = "MEMORÁNDUM DE FINALIZACIÓN DE RELACIÓN LABORAL"
        ws['A1'].font = Font(size=12, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
    
    def _fill_memo_data(self, ws, result: FiniquitoCalculationResult, include_cite: bool, cite_number: Optional[str]):
        """
        Fill memo with data
        """
        # Date and cite
        ws['A3'] = f"Fecha: {datetime.now().strftime('%d de %B de %Y')}"
        if include_cite and cite_number:
            ws['A4'] = f"CITE: {cite_number}"
        
        # Content
        row = 6
        ws[f'A{row}'] = "DE: Recursos Humanos"
        row += 1
        ws[f'A{row}'] = f"PARA: {result.employee.nombre}"
        row += 2
        
        ws[f'A{row}'] = f"Por medio del presente, se comunica la finalización de la relación laboral"
        row += 1
        ws[f'A{row}'] = f"con fecha {result.case_params.pay_until_date.strftime('%d/%m/%Y')}."
        row += 2
        
        ws[f'A{row}'] = f"Motivo: {result.case_params.motivo_retiro}"
        row += 2
        
        ws[f'A{row}'] = f"Se adjunta el cálculo de beneficios sociales correspondientes."
        
        # Save the workbook
        wb.save(output_path)
        return output_path
    
    def create_f_salida(self, calculation_result: FiniquitoCalculationResult,
                        template_path: Optional[str], output_path: str) -> str:
        """Create F-Salida document"""
        if template_path and os.path.exists(template_path):
            wb = load_workbook(template_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "F-Salida"
            self._create_f_salida_structure(ws)
        
        # Fill data
        self._fill_f_salida_data(ws, calculation_result)
        
        wb.save(output_path)
        return output_path
    
    def create_f_equipos(self, calculation_result: FiniquitoCalculationResult,
                         template_path: Optional[str], output_path: str) -> str:
        """Create F-Equipos document"""
        if template_path and os.path.exists(template_path):
            wb = load_workbook(template_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "F-Equipos"
            self._create_f_equipos_structure(ws)
        
        # Fill data
        self._fill_f_equipos_data(ws, calculation_result)
        
        wb.save(output_path)
        return output_path
    
    def create_contable_preview(self, calculation_result: FiniquitoCalculationResult,
                                template_path: Optional[str], output_path: str) -> str:
        """Create contable preview document (without account codes)"""
        if template_path and os.path.exists(template_path):
            wb = load_workbook(template_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Vista Contable"
            self._create_contable_structure(ws)
        
        # Fill data
        self._fill_contable_data(ws, calculation_result)
        
        wb.save(output_path)
        return output_path
    
    def create_rechazo_post(self, calculation_result: FiniquitoCalculationResult,
                           template_path: Optional[str], rejection_date: Optional[datetime],
                           output_path: str) -> str:
        """Create rechazo post-examen document"""
        if template_path and os.path.exists(template_path):
            wb = load_workbook(template_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Rechazo Post-Examen"
            self._create_rechazo_structure(ws)
        
        # Fill data
        self._fill_rechazo_data(ws, calculation_result, rejection_date)
        
        wb.save(output_path)
        return output_path
    
    def _create_f_salida_structure(self, ws):
        """Create structure for F-Salida"""
        ws['A1'] = "FORMULARIO DE SALIDA"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = "Datos del Empleado"
        ws['A3'].font = Font(bold=True, size=12)
        
        ws['A5'] = "CI:"
        ws['A6'] = "Nombre:"
        ws['A7'] = "Empresa:"
        ws['A8'] = "Unidad:"
        ws['A9'] = "Cargo:"
        ws['A10'] = "Fecha de Ingreso:"
        ws['A11'] = "Fecha de Salida:"
        ws['A12'] = "Motivo de Retiro:"
        
        ws['A14'] = "Entrega de Documentación"
        ws['A14'].font = Font(bold=True, size=12)
        
        ws['A16'] = "☐ Memorándum de designación"
        ws['A17'] = "☐ Contrato de trabajo"
        ws['A18'] = "☐ Credencial de trabajo"
        ws['A19'] = "☐ Tarjetas de presentación"
        ws['A20'] = "☐ Otros documentos institucionales"
        
        ws['A22'] = "Observaciones:"
        ws['A23'] = "_" * 60
        
        ws['A26'] = "Firma del Empleado"
        ws['A27'] = "_" * 30
        
        ws['D26'] = "Firma RRHH"
        ws['D27'] = "_" * 30
    
    def _fill_f_salida_data(self, ws, result: FiniquitoCalculationResult):
        """Fill F-Salida with data"""
        ws['B5'] = result.employee.ci
        ws['B6'] = result.employee.nombre
        ws['B7'] = result.employee.empresa
        ws['B8'] = result.employee.unidad
        ws['B9'] = result.employee.ocupacion
        ws['B10'] = result.employee.fecha_ingreso.strftime('%d/%m/%Y')
        ws['B11'] = result.case_params.pay_until_date.strftime('%d/%m/%Y')
        ws['B12'] = result.case_params.motivo_retiro
    
    def _create_f_equipos_structure(self, ws):
        """Create structure for F-Equipos"""
        ws['A1'] = "FORMULARIO DE DEVOLUCIÓN DE EQUIPOS"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = "Datos del Empleado"
        ws['A3'].font = Font(bold=True, size=12)
        
        ws['A5'] = "CI:"
        ws['A6'] = "Nombre:"
        ws['A7'] = "Unidad:"
        ws['A8'] = "Fecha de Devolución:"
        
        ws['A10'] = "Equipos a Devolver"
        ws['A10'].font = Font(bold=True, size=12)
        
        ws['A12'] = "☐ Computadora portátil"
        ws['A13'] = "☐ Teléfono móvil"
        ws['A14'] = "☐ Tarjetas de acceso"
        ws['A15'] = "☐ Llaves de oficina"
        ws['A16'] = "☐ Uniforme de trabajo"
        ws['A17'] = "☐ Herramientas de trabajo"
        ws['A18'] = "☐ Vehículo asignado"
        ws['A19'] = "☐ Otros: ________________"
        
        ws['A21'] = "Estado de los Equipos"
        ws['A21'].font = Font(bold=True, size=12)
        
        ws['A23'] = "Observaciones:"
        ws['A24'] = "_" * 60
        
        ws['A27'] = "Firma del Empleado"
        ws['A28'] = "_" * 30
        
        ws['D27'] = "Firma Sistemas/Activos"
        ws['D28'] = "_" * 30
    
    def _fill_f_equipos_data(self, ws, result: FiniquitoCalculationResult):
        """Fill F-Equipos with data"""
        ws['B5'] = result.employee.ci
        ws['B6'] = result.employee.nombre
        ws['B7'] = result.employee.unidad
        ws['B8'] = result.case_params.pay_until_date.strftime('%d/%m/%Y')
    
    def _create_contable_structure(self, ws):
        """Create structure for contable preview"""
        ws['A1'] = "VISTA PREVIA CONTABLE - FINIQUITO"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = "Información del Caso"
        ws['A3'].font = Font(bold=True, size=12)
        
        ws['A5'] = "Empleado:"
        ws['A6'] = "CI:"
        ws['A7'] = "Empresa:"
        ws['A8'] = "Fecha de Cálculo:"
        
        ws['A10'] = "Detalle de Cuentas"
        ws['A10'].font = Font(bold=True, size=12)
        
        # Headers for accounting table
        ws['A12'] = "Concepto"
        ws['B12'] = "Debe"
        ws['C12'] = "Haber"
        ws['D12'] = "Observaciones"
        
        for col in ['A12', 'B12', 'C12', 'D12']:
            ws[col].font = Font(bold=True)
    
    def _fill_contable_data(self, ws, result: FiniquitoCalculationResult):
        """Fill contable preview with data"""
        ws['B5'] = result.employee.nombre
        ws['B6'] = result.employee.ci
        ws['B7'] = result.employee.empresa
        ws['B8'] = datetime.now().strftime('%d/%m/%Y')
        
        current_row = 13
        
        # Add benefits as Debe (debit)
        for benefit in result.benefits.benefits_detail:
            ws[f'A{current_row}'] = benefit['concept']
            ws[f'B{current_row}'] = f"{benefit['amount']:,.2f}"
            ws[f'D{current_row}'] = benefit['description']
            current_row += 1
        
        # Add deductions as Haber (credit)
        for deduction in result.benefits.deductions_detail:
            ws[f'A{current_row}'] = deduction['concept']
            ws[f'C{current_row}'] = f"{deduction['amount']:,.2f}"
            current_row += 1
        
        # Add totals
        current_row += 1
        ws[f'A{current_row}'] = "TOTALES"
        ws[f'B{current_row}'] = f"{result.benefits.total_benefits:,.2f}"
        ws[f'C{current_row}'] = f"{result.benefits.total_deductions:,.2f}"
        
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'].font = Font(bold=True)
        ws[f'C{current_row}'].font = Font(bold=True)
        
        # Add net payment
        current_row += 2
        ws[f'A{current_row}'] = "NETO A PAGAR"
        ws[f'B{current_row}'] = f"{result.benefits.net_payment:,.2f}"
        ws[f'A{current_row}'].font = Font(bold=True, size=12)
        ws[f'B{current_row}'].font = Font(bold=True, size=12)
    
    def _create_rechazo_structure(self, ws):
        """Create structure for rechazo post-examen"""
        ws['A1'] = "NOTIFICACIÓN DE RECHAZO - EXAMEN POST-OCUPACIONAL"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = "Datos del Empleado"
        ws['A3'].font = Font(bold=True, size=12)
        
        ws['A5'] = "CI:"
        ws['A6'] = "Nombre:"
        ws['A7'] = "Empresa:"
        ws['A8'] = "Cargo:"
        ws['A9'] = "Fecha de Retiro:"
        ws['A10'] = "Fecha de Examen:"
        ws['A11'] = "Resultado:"
        
        ws['A13'] = "Detalles del Finiquito"
        ws['A13'].font = Font(bold=True, size=12)
        
        ws['A15'] = "Debido al resultado no apto en el examen post-ocupacional,"
        ws['A16'] = "el finiquito se procesa con las siguientes consideraciones:"
        
        ws['A18'] = "Total Beneficios:"
        ws['A19'] = "Total Deducciones:"
        ws['A20'] = "Neto a Pagar:"
        
        ws['A22'] = "Observaciones:"
        ws['A23'] = "_" * 60
        
        ws['A26'] = "Firma del Empleado"
        ws['A27'] = "_" * 30
        
        ws['D26'] = "Firma RRHH"
        ws['D27'] = "_" * 30
    
    def _fill_rechazo_data(self, ws, result: FiniquitoCalculationResult, rejection_date: Optional[datetime]):
        """Fill rechazo post-examen with data"""
        ws['B5'] = result.employee.ci
        ws['B6'] = result.employee.nombre
        ws['B7'] = result.employee.empresa
        ws['B8'] = result.employee.ocupacion
        ws['B9'] = result.case_params.pay_until_date.strftime('%d/%m/%Y')
        ws['B10'] = rejection_date.strftime('%d/%m/%Y') if rejection_date else 'N/A'
        ws['B11'] = "NO APTO"
        ws['B11'].font = Font(bold=True, color="FF0000")
        
        ws['B18'] = f"Bs. {result.benefits.total_benefits:,.2f}"
        ws['B19'] = f"Bs. {result.benefits.total_deductions:,.2f}"
        ws['B20'] = f"Bs. {result.benefits.net_payment:,.2f}"
        ws['B20'].font = Font(bold=True)
